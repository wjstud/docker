from openpyxl import load_workbook
from openpyxl import Workbook


def combine():
    wb = load_workbook('courses.xlsx')
    students = wb.get_sheet_by_name('students')
    time = wb.get_sheet_by_name('time')
    dict1 = {}
    dict2 = {}
    for i in range(2,students.max_row+1):
        dict1.setdefault(students.cell(row=i, column=2).value, [students.cell(row=i, column=1).value,students.cell(row=i,column=2).value,students.cell(row=i,column=3).value])

    for i in range(2,time.max_row+1):
        dict2.setdefault(time.cell(row=i, column=2).value, time.cell(row=i,column=3).value)

    new_list = [['创建时间','课程名称','学习人数','学习时间']]
    for k,v in dict1.items():
         for i,j in dict2.items():
             if k == i:
                 v.append(j)
                 new_list.append(v)


    ws = wb.create_sheet(index=2,title="combine")

    for row in new_list:
        ws.append(row)

    wb.save("courses.xlsx")

def split():
    wb = load_workbook("courses.xlsx")
    students = wb.get_sheet_by_name("combine")
    year_list = set()
    for i in range(2, students.max_row + 1):
        year_list.add(students.cell(row=i, column=1).value.year)

    dict1 = {}

    for year in year_list:
        values = [['创建时间','课程名称','学习人数','学习时间']]
        for i in range(2, students.max_row + 1):

            if year == students.cell(row=i, column=1).value.year:

                values.append(([students.cell(row=i, column=1).value, students.cell(row=i, column=2).value, students.cell(row=i, column=3).value, students.cell(row=i, column=4).value]))

        dict1[year] = values

    for k,v in dict1.items():
        newwb = Workbook()
        ws = newwb.active
        for year in year_list:
            if k == year:
                for i in v:
                    ws.append(i)
                newwb.save("%s.xlsx" % k)

if __name__ == '__main__':
    combine()
    split()
